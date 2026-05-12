import io
import csv
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, HRFlowable, Image
from reportlab.lib.styles import ParagraphStyle


def _get_company():
    try:
        from apps.users.models import CompanySettings
        return CompanySettings.get()
    except Exception:
        return None


def _company_header(story, styles):
    """Prepend company logo + name to a PDF story."""
    company = _get_company()
    name = company.name if company else 'Xelaltech221 CRM'

    header_data = [[]]
    # Logo cell
    if company and company.logo:
        try:
            img = Image(company.logo.path, width=2.5*cm, height=2.5*cm)
            img.hAlign = 'LEFT'
            header_data[0].append(img)
        except Exception:
            header_data[0].append('')
    else:
        header_data[0].append('')

    # Name cell
    title_style = ParagraphStyle('co', fontSize=18, textColor=colors.HexColor('#2563EB'),
                                  fontName='Helvetica-Bold', leading=22)
    sub_style = ParagraphStyle('sub', fontSize=9, textColor=colors.HexColor('#64748b'),
                                fontName='Helvetica')
    lines = [Paragraph(name, title_style)]
    if company and company.tagline:
        lines.append(Paragraph(company.tagline, sub_style))
    if company and company.email:
        lines.append(Paragraph(company.email, sub_style))
    if company and company.phone:
        lines.append(Paragraph(company.phone, sub_style))
    from reportlab.platypus import KeepTogether
    header_data[0].append(lines)

    t = Table(header_data, colWidths=[3*cm, 14*cm])
    t.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#2563EB'), spaceAfter=8))


def generate_report_file(report):
    """Generate a file (PDF/Excel/CSV) for the given Report and return (content_bytes, filename)."""
    fmt = report.export_format

    if fmt == 'pdf':
        return _report_pdf(report), f"{report.title}.pdf"
    elif fmt == 'excel':
        return _report_excel(report), f"{report.title}.xlsx"
    else:
        return _report_csv(report), f"{report.title}.csv"


def _report_pdf(report):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    _company_header(story, styles)
    story.append(Spacer(1, 0.3*cm))

    heading_style = ParagraphStyle('h', fontSize=14, textColor=colors.HexColor('#0f172a'),
                                    fontName='Helvetica-Bold', spaceAfter=8)
    story.append(Paragraph(report.title, heading_style))
    story.append(Spacer(1, 0.3*cm))

    meta = [
        ['Type:', report.get_report_type_display()],
        ['Période:', f"{report.period_start.strftime('%d/%m/%Y')} — {report.period_end.strftime('%d/%m/%Y')}"],
        ['Généré par:', str(report.generated_by) if report.generated_by else '—'],
    ]
    meta_table = Table(meta, colWidths=[4*cm, 12*cm])
    meta_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.5*cm))

    d = report.data or {}
    revenues  = d.get('revenues', 0)
    expenses  = d.get('expenses', 0)
    profit    = d.get('profit', revenues - expenses)
    txn_count = d.get('transaction_count', 0)

    # ── Section Transactions ──────────────────────────────────────
    section_style = ParagraphStyle('sec', fontSize=11, textColor=colors.HexColor('#0f172a'),
                                    fontName='Helvetica-Bold', spaceBefore=10, spaceAfter=4)
    story.append(Paragraph('Transactions', section_style))

    summary_data = [
        ['Indicateur', 'Montant'],
        ['Revenus', f"{revenues:,.0f} FCFA"],
        ['Dépenses', f"{expenses:,.0f} FCFA"],
        ['Bénéfice net', f"{profit:,.0f} FCFA"],
        ['Nombre de transactions', str(txn_count)],
    ]
    if 'by_category' in d:
        summary_data.append(['', ''])
        summary_data.append(['Catégorie', 'Montant dépensé'])
        for cat_name, amount in d['by_category'].items():
            summary_data.append([cat_name, f"{amount:,.0f} FCFA"])

    summary_table = Table(summary_data, colWidths=[9*cm, 7*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#EFF6FF')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)

    # ── Section Fonds de Caisse ───────────────────────────────────
    if any(k in d for k in ('caisse_opening', 'caisse_current')):
        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#E2E8F0'), spaceAfter=6))
        story.append(Paragraph('Fonds de Caisse', section_style))

        caisse_opening = d.get('caisse_opening', 0)
        caisse_credits = d.get('caisse_credits', 0)
        caisse_debits  = d.get('caisse_debits', 0)
        caisse_closing = d.get('caisse_closing', caisse_opening + caisse_credits - caisse_debits)
        caisse_current = d.get('caisse_current', caisse_closing)

        caisse_data = [
            ['', 'Montant'],
            ['Solde avant les transactions (ouverture)', f"{caisse_opening:,.0f} FCFA"],
            ['  + Entrées (alimentations)', f"+{caisse_credits:,.0f} FCFA"],
            ['  − Sorties (dépenses)', f"−{caisse_debits:,.0f} FCFA"],
            ['Solde de clôture période', f"{caisse_closing:,.0f} FCFA"],
            ['Solde actuel (après toutes transactions)', f"{caisse_current:,.0f} FCFA"],
        ]
        caisse_table = Table(caisse_data, colWidths=[9*cm, 7*cm])
        closing_color = colors.HexColor('#EFF6FF') if caisse_closing >= 0 else colors.HexColor('#FEF2F2')
        current_color = colors.HexColor('#DCFCE7') if caisse_current >= 0 else colors.HexColor('#FEF2F2')
        caisse_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
            ('FONTNAME', (0, 4), (-1, 5), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 4), (-1, 4), closing_color),
            ('BACKGROUND', (0, 5), (-1, 5), current_color),
            ('ROWBACKGROUNDS', (0, 1), (-1, 3), [colors.white, colors.HexColor('#F8FAFC')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 2), (0, 2), colors.HexColor('#16a34a')),
            ('TEXTCOLOR', (0, 3), (0, 3), colors.HexColor('#dc2626')),
        ]))
        story.append(caisse_table)

    doc.build(story)
    buf.seek(0)
    return buf.read()


def _report_excel(report):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport"

    blue = "2563EB"
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor=blue)
    center = Alignment(horizontal="center")

    ws.merge_cells("A1:B1")
    ws["A1"] = "FinTrack CRM — " + report.title
    ws["A1"].font = Font(bold=True, size=14, color=blue)

    ws["A3"] = "Type"
    ws["B3"] = report.get_report_type_display()
    ws["A4"] = "Période"
    ws["B4"] = f"{report.period_start.strftime('%d/%m/%Y')} — {report.period_end.strftime('%d/%m/%Y')}"

    ws["A6"] = "Indicateur"
    ws["B6"] = "Montant (FCFA)"
    for cell in [ws["A6"], ws["B6"]]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    d = report.data or {}
    rows = [
        ("Revenus", d.get('revenues', 0)),
        ("Dépenses", d.get('expenses', 0)),
        ("Bénéfice net", d.get('profit', 0)),
        ("Transactions", d.get('transaction_count', 0)),
    ]
    for i, (label, value) in enumerate(rows, start=7):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value

    row = 7 + len(rows)
    if 'by_category' in d:
        ws[f"A{row}"] = "Catégorie"
        ws[f"B{row}"] = "Montant (FCFA)"
        for cell in [ws[f"A{row}"], ws[f"B{row}"]]:
            cell.font = header_font
            cell.fill = header_fill
        row += 1
        for cat_name, amount in d['by_category'].items():
            ws[f"A{row}"] = cat_name
            ws[f"B{row}"] = amount
            row += 1

    # ── Section Fonds de Caisse ───────────────────────────────────
    if any(k in d for k in ('caisse_opening', 'caisse_current')):
        row += 1
        dark_fill = PatternFill("solid", fgColor="0F172A")
        green_fill = PatternFill("solid", fgColor="DCFCE7")
        red_fill   = PatternFill("solid", fgColor="FEF2F2")

        ws[f"A{row}"] = "FONDS DE CAISSE"
        ws[f"B{row}"] = "Montant (FCFA)"
        for cell in [ws[f"A{row}"], ws[f"B{row}"]]:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = dark_fill
        row += 1

        caisse_rows = [
            ("Solde avant les transactions (ouverture)", d.get('caisse_opening', 0)),
            ("+ Entrées (alimentations)",                d.get('caisse_credits', 0)),
            ("− Sorties (dépenses)",                    d.get('caisse_debits',  0)),
            ("Solde de clôture période",                 d.get('caisse_closing', 0)),
            ("Solde actuel (après toutes transactions)", d.get('caisse_current', 0)),
        ]
        for i, (label, value) in enumerate(caisse_rows):
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            if i == 3:
                f = green_fill if value >= 0 else red_fill
                ws[f"A{row}"].fill = f
                ws[f"B{row}"].fill = f
                ws[f"A{row}"].font = Font(bold=True, size=11)
                ws[f"B{row}"].font = Font(bold=True, size=11)
            if i == 4:
                f = green_fill if value >= 0 else red_fill
                ws[f"A{row}"].fill = f
                ws[f"B{row}"].fill = f
                ws[f"A{row}"].font = Font(bold=True, size=11, color="15803D" if value >= 0 else "DC2626")
                ws[f"B{row}"].font = Font(bold=True, size=11, color="15803D" if value >= 0 else "DC2626")
            row += 1

    for col in ['A', 'B']:
        ws.column_dimensions[col].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _report_csv(report):
    buf = io.StringIO()
    writer = csv.writer(buf)
    d = report.data or {}

    writer.writerow(["FinTrack CRM — " + report.title])
    writer.writerow(["Type", report.get_report_type_display()])
    writer.writerow(["Période", f"{report.period_start.strftime('%d/%m/%Y')} — {report.period_end.strftime('%d/%m/%Y')}"])
    writer.writerow([])
    writer.writerow(["Indicateur", "Montant"])
    writer.writerow(["Revenus", d.get('revenues', 0)])
    writer.writerow(["Dépenses", d.get('expenses', 0)])
    writer.writerow(["Bénéfice net", d.get('profit', 0)])
    writer.writerow(["Transactions", d.get('transaction_count', 0)])

    if 'by_category' in d:
        writer.writerow([])
        writer.writerow(["Catégorie", "Montant"])
        for cat_name, amount in d['by_category'].items():
            writer.writerow([cat_name, amount])

    if any(k in d for k in ('caisse_opening', 'caisse_current')):
        writer.writerow([])
        writer.writerow(["FONDS DE CAISSE", "Montant (FCFA)"])
        writer.writerow(["Solde avant les transactions (ouverture)", d.get('caisse_opening', 0)])
        writer.writerow(["+ Entrées (alimentations)",               d.get('caisse_credits', 0)])
        writer.writerow(["− Sorties (dépenses)",                   d.get('caisse_debits',  0)])
        writer.writerow(["Solde de clôture période",                d.get('caisse_closing', 0)])
        writer.writerow(["Solde actuel (après toutes transactions)", d.get('caisse_current', 0)])

    return buf.getvalue().encode('utf-8')


def generate_invoice_pdf(invoice):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                             rightMargin=2*cm, leftMargin=2*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Company header with logo
    _company_header(story, styles)
    story.append(Spacer(1, 0.3*cm))

    inv_style = ParagraphStyle('inv', fontSize=16, textColor=colors.HexColor('#0f172a'),
                                fontName='Helvetica-Bold', spaceAfter=6)
    story.append(Paragraph(f"FACTURE {invoice.invoice_number}", inv_style))
    story.append(Spacer(1, 0.3*cm))

    # Client info
    info_data = [
        ['Client:', invoice.client_name],
        ['Email:', invoice.client_email or '-'],
        ['Adresse:', invoice.client_address or '-'],
        ['Date émission:', invoice.issue_date.strftime('%d/%m/%Y')],
        ['Date échéance:', invoice.due_date.strftime('%d/%m/%Y')],
        ['Statut:', invoice.get_status_display()],
    ]
    info_table = Table(info_data, colWidths=[4*cm, 12*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.5*cm))

    # Items
    headers = ['Description', 'Qté', 'Prix unitaire', 'Total']
    rows = [headers]
    for item in invoice.items.all():
        rows.append([
            item.description,
            str(item.quantity),
            f"{item.unit_price:,.0f} FCFA",
            f"{item.total:,.0f} FCFA",
        ])
    rows.append(['', '', 'Sous-total:', f"{invoice.subtotal:,.0f} FCFA"])
    rows.append(['', '', f'TVA ({invoice.tax_rate}%):', f"{invoice.tax_amount:,.0f} FCFA"])
    rows.append(['', '', 'Remise:', f"-{invoice.discount:,.0f} FCFA"])
    rows.append(['', '', 'TOTAL:', f"{invoice.total:,.0f} FCFA"])

    items_table = Table(rows, colWidths=[8*cm, 2*cm, 4*cm, 3*cm])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (2, -3), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EFF6FF')),
        ('GRID', (0, 0), (-1, -4), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -4), [colors.white, colors.HexColor('#F8FAFC')]),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(items_table)

    if invoice.notes:
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(f"<b>Notes:</b> {invoice.notes}", styles['Normal']))

    doc.build(story)
    buffer.seek(0)
    return buffer
